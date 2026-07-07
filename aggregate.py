"""Excel/CSV aggregation skill for Lyzr agents.

Public entry point: :func:`aggregate_data`, which loads tabular data from a file
path or inline text, performs a dynamic group-by aggregation described by the
caller, and returns the result as a structured JSON string.

Register with a Lyzr agent (see ``register_example.py``)::

    from lyzr import Studio
    from aggregate import aggregate_data

    studio = Studio()
    agent = studio.create_agent(...)
    agent.add_tool(aggregate_data)
"""

from __future__ import annotations

import hashlib
import json
import math
import os
import re
from io import StringIO

import numpy as np
import pandas as pd


class AggregationError(Exception):
    """Raised for recoverable, user-facing errors. Surfaced as JSON, not a stack trace."""

    def __init__(self, error_type: str, message: str) -> None:
        super().__init__(message)
        self.error_type = error_type
        self.message = message


# User-facing aggregation function -> pandas aggregation name.
_FUNC_ALIASES = {
    "sum": "sum",
    "mean": "mean",
    "avg": "mean",
    "average": "mean",
    "min": "min",
    "max": "max",
    "count": "count",
    "nunique": "nunique",
    "count_distinct": "nunique",
    "distinct": "nunique",
    "median": "median",
    "std": "std",
    "var": "var",
    "first": "first",
    "last": "last",
}

# Functions that operate on numbers; their target columns are coerced to numeric.
_NUMERIC_FUNCS = {"sum", "mean", "median", "std", "var"}

# Stripped from formatted numeric strings (currency symbols + whitespace) before a
# second coercion attempt, so values like "$1,234.50" are recovered, not silently dropped.
_CURRENCY_WS_RE = r"[\s$€£¥₹]"
# A number written with thousands-separator commas, e.g. "1,234" or "-12,345.60".
_THOUSANDS_RE = r"^-?\d{1,3}(,\d{3})+(\.\d+)?$"

_CSV_EXTS = {".csv", ".txt"}
_TSV_EXTS = {".tsv"}
_EXCEL_EXTS = {".xlsx", ".xlsm", ".xls"}
# Extensions we can write results to. .xlsm is macro-enabled and needs a template
# because openpyxl can preserve VBA but cannot author it from scratch.
_EXCEL_OUTPUT_EXTS = {".xlsx", ".xlsm"}


# --------------------------------------------------------------------------- #
# Argument normalization (robust to the loose types an LLM may pass)
# --------------------------------------------------------------------------- #
def _normalize_str_list(value) -> list:
    """Accept a list/tuple, a JSON-array string, or a comma-separated string."""
    if value is None:
        return []
    if isinstance(value, (list, tuple)):
        return [str(item).strip() for item in value if str(item).strip()]
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return []
        if text[:1] == "[":
            try:
                parsed = json.loads(text)
                if isinstance(parsed, list):
                    return [str(item).strip() for item in parsed if str(item).strip()]
            except json.JSONDecodeError:
                pass
        return [part.strip() for part in text.split(",") if part.strip()]
    raise AggregationError(
        "BadArgument",
        f"Expected a list or string but got {type(value).__name__}.",
    )


def _to_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "y", "desc", "descending"}
    return bool(value)


def _to_int(value) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


# --------------------------------------------------------------------------- #
# Input loading
# --------------------------------------------------------------------------- #
def _resolve_sheet(sheet_name: str):
    text = (sheet_name or "").strip()
    if not text:
        return 0
    return int(text) if text.isdigit() else text


# Keys that commonly wrap a list of records in a JSON response.
_JSON_RECORD_KEYS = ("data", "records", "rows", "items", "results", "result", "values")


def _parse_json_text(text: str):
    """Parse a JSON document, falling back to JSON Lines (NDJSON) when needed."""
    text = text.lstrip("\ufeff").strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError as exc:
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        if len(lines) > 1:
            try:
                return [json.loads(line) for line in lines]
            except json.JSONDecodeError:
                pass
        raise AggregationError("ParseError", f"Invalid JSON: {exc.msg}") from exc


def _is_dataset_like(obj) -> bool:
    """True for a dict describing one table via parallel ``headers`` and ``rows`` lists."""
    return (
        isinstance(obj, dict)
        and isinstance(obj.get("headers"), list)
        and isinstance(obj.get("rows"), list)
    )


def _all_dataset_like(value) -> bool:
    """True for a non-empty list whose every item is :func:`_is_dataset_like`."""
    return (
        isinstance(value, list)
        and len(value) > 0
        and all(_is_dataset_like(item) for item in value)
    )


def _find_datasets(obj) -> list:
    """Locate a list of ``headers``/``rows`` datasets anywhere in a parsed JSON value.

    Recognizes a single dataset object, a bare list of dataset objects, an explicit
    ``{"datasets": [...]}`` container, and datasets nested under wrapper keys such as
    ``{"response": {"result": {"datasets": [...]}}}`` (e.g. another agent's output).
    Returns an empty list for record-oriented JSON, which is handled elsewhere. Only dict
    values are descended into; list values are cheaply tested (record lists fail fast on
    their first item), so large record arrays are not deeply scanned.
    """
    if _is_dataset_like(obj):
        return [obj]
    if _all_dataset_like(obj):
        return list(obj)
    if isinstance(obj, dict):
        if _all_dataset_like(obj.get("datasets")):
            return list(obj["datasets"])
        for value in obj.values():
            if isinstance(value, dict):
                found = _find_datasets(value)
                if found:
                    return found
            elif _all_dataset_like(value):
                return list(value)
    return []


def _select_dataset(datasets, dataset):
    """Pick one dataset by name or 0-based index; return ``(dataset, name)``.

    With a single dataset, selection is optional. With several, ``dataset`` must name one
    (case-insensitive) or give its index; otherwise an ``AmbiguousDataset`` error lists the
    available names so the caller can retry with a specific one.
    """
    names = [str(item.get("name") or index) for index, item in enumerate(datasets)]
    selection = (dataset or "").strip()

    if not selection:
        if len(datasets) == 1:
            return datasets[0], names[0]
        raise AggregationError(
            "AmbiguousDataset",
            f"The input contains {len(datasets)} datasets ({names}). Select one with "
            "'dataset' (a dataset name or a 0-based index).",
        )

    for item, name in zip(datasets, names):
        if name.lower() == selection.lower():
            return item, name
    if selection.isdigit():
        index = int(selection)
        if 0 <= index < len(datasets):
            return datasets[index], names[index]
    raise AggregationError(
        "DatasetNotFound",
        f"Dataset '{selection}' not found. Available datasets: {names}.",
    )


def _frame_from_dataset(dataset, name):
    """Turn one ``{"headers": [...], "rows": [...]}`` dataset into a DataFrame.

    ``rows`` is normally a list of value lists aligned to ``headers``; a list of row
    objects (dicts) is also tolerated and re-ordered to the declared header order.
    """
    headers = dataset.get("headers")
    rows = dataset.get("rows")
    if not isinstance(headers, list) or not isinstance(rows, list):
        raise AggregationError(
            "ParseError", f"Dataset '{name}' must provide list 'headers' and 'rows'."
        )
    columns = [str(header) for header in headers]

    if rows and all(isinstance(row, dict) for row in rows):
        frame = pd.json_normalize(rows)
        ordered = [column for column in columns if column in frame.columns]
        extra = [column for column in frame.columns if column not in ordered]
        return frame[ordered + extra]

    for index, row in enumerate(rows):
        if not isinstance(row, (list, tuple)):
            raise AggregationError(
                "ParseError", f"Dataset '{name}' row {index} is not a list of values."
            )
        if len(row) != len(columns):
            raise AggregationError(
                "ParseError",
                f"Dataset '{name}' row {index} has {len(row)} values but there are "
                f"{len(columns)} headers.",
            )
    return pd.DataFrame(rows, columns=columns)


def _frame_from_records(obj):
    """Build a DataFrame from a record-oriented parsed JSON value.

    Handles a list of record objects, a single record object, records wrapped under a
    key (e.g. {"data": [...]}), column-oriented data ({"col": [...], ...}), and nested
    objects (flattened into dotted column names).
    """
    if isinstance(obj, dict):
        record_list = None
        for key in _JSON_RECORD_KEYS:
            candidate = obj.get(key)
            if (
                isinstance(candidate, list)
                and candidate
                and all(isinstance(item, dict) for item in candidate)
            ):
                record_list = candidate
                break
        if record_list is None:
            list_keys = [key for key, value in obj.items() if isinstance(value, list)]
            if len(list_keys) == 1:
                candidate = obj[list_keys[0]]
                if candidate and all(isinstance(item, dict) for item in candidate):
                    record_list = candidate
        if record_list is not None:
            obj = record_list
        else:
            values = list(obj.values())
            if values and all(isinstance(value, list) for value in values):
                return pd.DataFrame(obj)  # column-oriented data
            obj = [obj]  # a single record object

    if isinstance(obj, list):
        if obj and all(isinstance(item, dict) for item in obj):
            return pd.json_normalize(obj)
        return pd.DataFrame({"value": obj})

    return pd.DataFrame({"value": [obj]})


def _frame_from_json(obj, dataset=""):
    """Build a DataFrame from a parsed JSON value in any common shape.

    Returns a ``(DataFrame, dataset_name)`` pair. ``dataset_name`` is the selected table's
    name when the input is a multi-table "datasets" envelope (each table given as
    ``headers`` + ``rows``, optionally nested under wrapper keys); otherwise it is ``None``.
    Record-oriented shapes (list of objects, single object, wrapped-under-key,
    column-oriented, nested, NDJSON) are delegated to :func:`_frame_from_records`.
    """
    datasets = _find_datasets(obj)
    if datasets:
        chosen, name = _select_dataset(datasets, dataset)
        return _frame_from_dataset(chosen, name), name
    return _frame_from_records(obj), None


def _load_from_file(file_path: str, data_format: str, sheet_name: str, dataset: str):
    if not os.path.isfile(file_path):
        raise AggregationError("FileNotFound", f"File not found: {file_path}")

    ext = os.path.splitext(file_path)[1].lower()
    resolved = data_format
    if resolved == "auto":
        if ext in _EXCEL_EXTS:
            resolved = "excel"
        elif ext == ".json":
            resolved = "json"
        elif ext in _TSV_EXTS:
            resolved = "tsv"
        else:
            resolved = "csv"

    dataset_name = None
    try:
        if resolved == "excel":
            frame = pd.read_excel(file_path, sheet_name=_resolve_sheet(sheet_name))
        elif resolved == "json":
            with open(file_path, "r", encoding="utf-8-sig") as handle:
                frame, dataset_name = _frame_from_json(_parse_json_text(handle.read()), dataset)
        elif resolved == "tsv":
            frame = pd.read_csv(file_path, sep="\t", encoding="utf-8-sig")
        else:
            frame = pd.read_csv(file_path, encoding="utf-8-sig")
    except AggregationError:
        raise
    except Exception as exc:  # noqa: BLE001 - report parse issues as JSON
        raise AggregationError("ParseError", f"Failed to read '{file_path}': {exc}") from exc

    source = {
        "type": "file",
        "location": file_path,
        "format": "csv" if resolved == "tsv" else resolved,
        "sheet": sheet_name or None if resolved == "excel" else None,
        "dataset": dataset_name,
    }
    return frame, source


def _load_from_inline(data: str, data_format: str, dataset: str):
    text = data.lstrip("\ufeff").strip()
    resolved = data_format
    if resolved in {"auto", "excel"}:
        resolved = "json" if text[:1] in {"[", "{"} else "csv"

    dataset_name = None
    try:
        if resolved == "json":
            frame, dataset_name = _frame_from_json(_parse_json_text(text), dataset)
        elif resolved == "tsv":
            frame = pd.read_csv(StringIO(text), sep="\t")
        else:
            frame = pd.read_csv(StringIO(text))
    except AggregationError:
        raise
    except Exception as exc:  # noqa: BLE001 - report parse issues as JSON
        raise AggregationError(
            "ParseError", f"Failed to parse inline data as {resolved}: {exc}"
        ) from exc

    source = {
        "type": "inline",
        "location": None,
        "format": resolved,
        "sheet": None,
        "dataset": dataset_name,
    }
    return frame, source


def _load_dataframe(file_path: str, data: str, data_format: str, sheet_name: str, dataset: str):
    file_path = (file_path or "").strip()
    data = data or ""
    data_format = (data_format or "auto").strip().lower()

    if file_path:
        return _load_from_file(file_path, data_format, sheet_name, dataset)
    if data.strip():
        return _load_from_inline(data, data_format, dataset)
    raise AggregationError(
        "NoInput", "No input provided. Set either 'file_path' or 'data'."
    )


# --------------------------------------------------------------------------- #
# Metric parsing and aggregation
# --------------------------------------------------------------------------- #
def _parse_metrics(metrics: list, columns) -> list:
    if not metrics:
        raise AggregationError(
            "NoMetrics", "Provide at least one metric, e.g. 'sum:amount' or 'count:*'."
        )

    columns = list(columns)
    specs = []
    used_aliases: set = set()

    for raw in metrics:
        parts = [part.strip() for part in str(raw).split(":")]
        func_key = parts[0].lower()
        column = parts[1] if len(parts) > 1 else "*"
        custom_alias = parts[2] if len(parts) > 2 and parts[2] else None

        if func_key not in _FUNC_ALIASES:
            raise AggregationError(
                "UnknownFunction",
                f"Unknown function '{parts[0]}' in metric '{raw}'. "
                f"Supported: {sorted(set(_FUNC_ALIASES))}.",
            )

        is_count_star = column in {"", "*"}
        if is_count_star:
            if func_key != "count":
                raise AggregationError(
                    "BadMetric",
                    f"Metric '{raw}' targets all rows ('*') but only 'count' is valid there.",
                )
            column = "*"
        elif column not in columns:
            raise AggregationError(
                "InvalidColumn",
                f"Column '{column}' in metric '{raw}' not found. Available columns: {columns}.",
            )

        alias = custom_alias or ("count" if is_count_star else f"{column}_{func_key}")
        base_alias = alias
        suffix = 2
        while alias in used_aliases:
            alias = f"{base_alias}_{suffix}"
            suffix += 1
        used_aliases.add(alias)

        specs.append(
            {
                "raw": str(raw),
                "func_key": func_key,
                "func": _FUNC_ALIASES[func_key],
                "column": column,
                "alias": alias,
                "is_count_star": is_count_star,
            }
        )
    return specs


def _coerce_numeric(series):
    """Coerce a column to numeric, tolerating common formatted-number strings.

    Values are parsed directly first; anything that fails (because an upstream step
    emitted it as text) is retried after stripping currency symbols and whitespace,
    turning accounting-style parentheses into a leading minus (``(1,234)`` -> ``-1234``),
    and removing thousands-separator commas -- but only when the commas group digits in
    threes, so a decimal comma such as ``"1,5"`` is left unparsed rather than silently
    misread as ``15``.

    Returns ``(numeric_series, unparsed_values)``. ``unparsed_values`` holds the originally
    non-null cells that still could not be parsed and became NaN (and were therefore
    excluded from sum/mean/median/std/var). Genuine nulls are not reported.
    """
    numeric = pd.to_numeric(series, errors="coerce")
    failed = series.notna() & numeric.isna()
    if failed.any():
        text = series[failed].astype(str).str.strip()
        text = text.str.replace(_CURRENCY_WS_RE, "", regex=True)
        text = text.str.replace(r"^\((.+)\)$", r"-\1", regex=True)
        is_thousands = text.str.match(_THOUSANDS_RE).fillna(False)
        text = text.mask(is_thousands, text.str.replace(",", "", regex=False))
        numeric.loc[failed] = pd.to_numeric(text, errors="coerce")
    unparsed = series[series.notna() & numeric.isna()]
    return numeric, unparsed


def _strip_group_keys(frame, group_by):
    """Trim leading/trailing whitespace from string values in group-by columns.

    Grouping is exact, so ``" 1"``, ``"1 "`` and ``"1"`` would otherwise form three
    different groups. Only text cells are trimmed; numbers, dates, and nulls are left
    untouched (so mixed-type columns are safe). Returns the frame unchanged when there is
    nothing to trim, otherwise a copy with the affected group-by columns trimmed.
    """
    to_strip = []
    for column in group_by:
        series = frame[column]
        if not (
            pd.api.types.is_object_dtype(series) or pd.api.types.is_string_dtype(series)
        ):
            continue  # numeric/datetime keys cannot carry surrounding whitespace
        changed = False
        values = []
        for value in series.tolist():
            if isinstance(value, str):
                trimmed = value.strip()
                changed = changed or trimmed != value
                values.append(trimmed)
            else:
                values.append(value)
        if changed:
            to_strip.append((column, pd.Series(values, index=series.index)))
    if not to_strip:
        return frame
    work = frame.copy()
    for column, stripped in to_strip:
        work[column] = stripped
    return work


def _group_key_warnings(frame, group_by) -> list:
    """Warn when a group-by column's values would collapse after light normalization.

    Leading/trailing whitespace is already trimmed before grouping (see
    :func:`_strip_group_keys`), so this flags the differences that are *not* auto-resolved:
    values that differ only by type (``"1"`` text vs ``1`` number) or by letter case
    (``"North"`` vs ``"north"``). Because grouping is exact, those still fragment one logical
    group into several -- so each fragment's total looks too low. This does not change the
    result; it flags the columns where fragmentation is happening so the caller can normalize
    the source (or trust the split intentionally).
    """
    warnings = []
    for column in group_by:
        series = frame[column]
        present = series[series.notna()]
        if present.empty:
            continue
        raw_distinct = present.unique()
        normalized = present.map(lambda value: str(value).strip().casefold())
        buckets: dict = {}
        for raw, norm in zip(present.tolist(), normalized.tolist()):
            buckets.setdefault(norm, set()).add(repr(raw))
        collisions = {norm: reprs for norm, reprs in buckets.items() if len(reprs) > 1}
        if collisions and len(normalized.unique()) < len(raw_distinct):
            examples = [sorted(reprs) for reprs in list(collisions.values())[:3]]
            warnings.append(
                {
                    "type": "inconsistent_group_keys",
                    "column": column,
                    "distinct_before": int(len(raw_distinct)),
                    "distinct_after_normalization": int(len(normalized.unique())),
                    "examples": examples,
                    "message": (
                        f"Group-by column '{column}' has values that differ only by type "
                        f"or letter case and were split into separate groups (e.g. "
                        f"{examples[0]}). Totals per group may look too low. Normalize the "
                        f"column upstream for a stable grouping."
                    ),
                }
            )
    return warnings


# --------------------------------------------------------------------------- #
# Input fingerprint (drift diagnostics)
# --------------------------------------------------------------------------- #
# A canonical token for any null/NaN cell, so nulls compare equal across runs and formats.
_NULL_TOKEN = "\x00null"


def _canonical_cell(value) -> str:
    """Return a representation-stable token for one cell value.

    Equivalent values collapse to the same token so cosmetic differences do not change a
    digest: integral floats and ints share ``num:<int>`` (``2000`` and ``2000.0`` ->
    ``num:2000``), all nulls/NaN/inf become ``_NULL_TOKEN``, datetimes use ISO format.
    Because numeric metric columns are fingerprinted *after* coercion and group-by columns
    *after* whitespace trimming, a differing digest means the effective aggregated values
    changed -- i.e. the input data drifted -- not merely how it was formatted.
    """
    if value is None:
        return _NULL_TOKEN
    try:
        if pd.isna(value):
            return _NULL_TOKEN
    except (TypeError, ValueError):
        pass
    if isinstance(value, np.generic):
        value = value.item()
    if isinstance(value, bool):  # must precede int: bool is a subclass of int
        return "bool:1" if value else "bool:0"
    if isinstance(value, int):
        return f"num:{value}"
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return _NULL_TOKEN
        if value.is_integer():
            return f"num:{int(value)}"
        return f"num:{value!r}"
    if hasattr(value, "isoformat"):
        return f"dt:{value.isoformat()}"
    return f"str:{value}"


def _digest_tokens(tokens) -> str:
    """Hash an iterable of string tokens into a short, stable hex digest."""
    hasher = hashlib.sha256()
    for token in tokens:
        hasher.update(token.encode("utf-8"))
        hasher.update(b"\x1e")  # record separator keeps token boundaries unambiguous
    return hasher.hexdigest()[:16]


def _column_fingerprint(series) -> dict:
    """Order-insensitive digest of one column's values, with null accounting.

    Values are treated as a multiset (sorted, duplicates kept), so row order does not
    affect the digest but any changed/added/removed value does. Lets a caller pinpoint
    which column changed between two runs that produced different totals.
    """
    tokens = [_canonical_cell(value) for value in series.tolist()]
    non_null = sum(1 for token in tokens if token != _NULL_TOKEN)
    return {
        "digest": _digest_tokens(sorted(tokens)),
        "non_null": non_null,
        "null": len(tokens) - non_null,
    }


def _input_fingerprint(work, group_by, specs) -> dict:
    """Build a drift-diagnostic fingerprint of the values feeding the aggregation.

    Fingerprints only the columns that can affect the result -- the group-by keys and the
    metric source columns -- using their *effective* values (post whitespace-trim and
    numeric coercion, exactly as ``work`` holds them). ``digest`` is an order-insensitive
    hash over those columns' rows (alignment within a row preserved, duplicates kept), so
    two runs with identical data produce an identical digest regardless of row order, while
    any changed value flips it. Per-column digests localize the change. When the only metric
    is ``count:*`` (no source columns) the whole frame is fingerprinted so ``digest`` still
    reflects the data.
    """
    metric_columns = [spec["column"] for spec in specs if not spec["is_count_star"]]
    involved = list(dict.fromkeys([*group_by, *metric_columns]))

    subject = work[involved] if involved else work
    row_tokens = [
        "\x1f".join(_canonical_cell(value) for value in row)
        for row in subject.itertuples(index=False, name=None)
    ]
    return {
        "rows": int(len(work)),
        "digest": _digest_tokens(sorted(row_tokens)),
        "columns": {
            str(column): _column_fingerprint(work[column]) for column in involved
        },
    }


def _run_aggregation(frame, group_by, specs, sort_by, descending, limit, include_fingerprint=False):
    columns = list(frame.columns)
    for column in group_by:
        if column not in columns:
            raise AggregationError(
                "InvalidColumn",
                f"Group-by column '{column}' not found. Available columns: {columns}.",
            )

    # Duplicate header names (e.g. produced by an upstream column-normalization step) make a
    # column reference ambiguous: frame[name] returns a 2-D frame, not a Series, which would
    # otherwise fail deep inside pandas with a cryptic message. Fail early and clearly for
    # any column actually referenced by group_by or a metric.
    referenced = set(group_by) | {
        spec["column"] for spec in specs if not spec["is_count_star"]
    }
    duplicated = sorted(
        {str(name) for name in columns if list(columns).count(name) > 1} & referenced
    )
    if duplicated:
        raise AggregationError(
            "DuplicateColumn",
            f"The data has duplicate column name(s) {duplicated}, so the aggregation is "
            "ambiguous. Rename or drop the duplicate column(s) before aggregating.",
        )

    numeric_aliases: dict = {}
    for spec in specs:
        if spec["func_key"] in _NUMERIC_FUNCS and not spec["is_count_star"]:
            numeric_aliases.setdefault(spec["column"], []).append(spec["alias"])

    work = frame
    warnings: list = []
    if group_by:
        work = _strip_group_keys(work, group_by)
    warnings.extend(_group_key_warnings(work, group_by))
    if numeric_aliases:
        if work is frame:
            work = frame.copy()
        for column, aliases in numeric_aliases.items():
            coerced, unparsed = _coerce_numeric(work[column])
            work[column] = coerced
            if int(coerced.notna().sum()) == 0:
                # No numeric values at all -> the metric is empty/zero. This most often
                # means the wrong dataset or column was selected (e.g. summing a money
                # column that is only populated in a different dataset), so flag it loudly
                # rather than returning a silent 0.
                samples = list(dict.fromkeys(str(value) for value in unparsed.tolist()))[:3]
                warning = {
                    "type": "no_numeric_values",
                    "column": column,
                    "affected_metrics": aliases,
                    "row_count": int(len(coerced)),
                    "message": (
                        f"Column '{column}' has no numeric values across {len(coerced)} "
                        f"row(s), so {', '.join(aliases)} is empty/zero. This often means the "
                        f"wrong dataset or column was selected."
                    ),
                }
                if samples:
                    warning["sample_values"] = samples
                warnings.append(warning)
            elif len(unparsed):
                samples = list(dict.fromkeys(str(value) for value in unparsed.tolist()))[:3]
                warnings.append(
                    {
                        "type": "non_numeric_values_excluded",
                        "column": column,
                        "affected_metrics": aliases,
                        "excluded_count": int(len(unparsed)),
                        "sample_values": samples,
                        "message": (
                            f"{len(unparsed)} value(s) in column '{column}' are not numeric "
                            f"and were excluded from {', '.join(aliases)}. Examples: {samples}."
                        ),
                    }
                )

    fingerprint = (
        _input_fingerprint(work, group_by, specs) if include_fingerprint else None
    )

    if group_by:
        grouped = work.groupby(group_by, dropna=False, sort=True)
        series_list = []
        aliases = []
        for spec in specs:
            aliases.append(spec["alias"])
            if spec["is_count_star"]:
                series_list.append(grouped.size())
            else:
                series_list.append(grouped[spec["column"]].agg(spec["func"]))
        result = pd.concat(series_list, axis=1, keys=aliases).reset_index()
    else:
        row = {}
        for spec in specs:
            if spec["is_count_star"]:
                row[spec["alias"]] = int(len(work))
            else:
                row[spec["alias"]] = work[spec["column"]].agg(spec["func"])
        result = pd.DataFrame([row])

    if sort_by:
        if sort_by not in result.columns:
            raise AggregationError(
                "InvalidSort",
                f"sort_by '{sort_by}' is not an output column. "
                f"Available: {list(result.columns)}.",
            )
        result = result.sort_values(
            by=sort_by, ascending=not descending, kind="mergesort"
        )

    if limit and limit > 0:
        result = result.head(limit)

    return result, warnings, fingerprint


# --------------------------------------------------------------------------- #
# JSON-safe output building
# --------------------------------------------------------------------------- #
def _json_safe(value):
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, np.generic):
        value = value.item()
    if hasattr(value, "isoformat"):
        return value.isoformat()
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    return value


def _records_from_frame(result) -> list:
    raw_records = result.to_dict(orient="records")
    return [{key: _json_safe(val) for key, val in record.items()} for record in raw_records]


def _build_output(
    frame, source, group_by, metrics, result, output_info=None, warnings=None,
    fingerprint=None,
) -> dict:
    records = _records_from_frame(result)
    output = {
        "status": "success",
        "source": source,
        "rows_read": int(len(frame)),
        "columns": [str(column) for column in frame.columns],
        "group_by": list(group_by),
        "metrics": list(metrics),
        "record_count": len(records),
        "records": records,
    }
    if warnings:
        output["warnings"] = warnings
    if fingerprint is not None:
        output["input_fingerprint"] = fingerprint
    if output_info is not None:
        output["output_file"] = output_info
    return output


# --------------------------------------------------------------------------- #
# Excel output (.xlsx / macro-preserving .xlsm)
# --------------------------------------------------------------------------- #
def _excel_safe(value):
    """Coerce a single value into something openpyxl can write to a cell."""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, np.generic):
        value = value.item()
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    # openpyxl refuses timezone-aware datetimes; drop the tz (or stringify if that fails).
    if getattr(value, "tzinfo", None) is not None:
        try:
            value = value.replace(tzinfo=None)
        except (TypeError, ValueError):
            value = str(value)
    return value


def _write_results_file(result, output_path, template_path, output_sheet) -> dict:
    """Write the aggregated ``result`` to an .xlsx/.xlsm file; return output metadata.

    A macro-enabled ``.xlsm`` requires ``template_path`` pointing at an existing
    macro-enabled workbook: openpyxl can *preserve* its VBA (``keep_vba=True``) but
    cannot create macros from scratch. The target ``output_sheet`` is (re)created with
    a header row plus one row per aggregated record; any other sheets in the template
    are left untouched.
    """
    output_path = output_path.strip()
    template_path = (template_path or "").strip()
    ext = os.path.splitext(output_path)[1].lower()
    if ext not in _EXCEL_OUTPUT_EXTS:
        raise AggregationError(
            "UnsupportedOutput",
            f"output_path must end in .xlsx or .xlsm, but got '{ext or output_path}'.",
        )

    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:  # pragma: no cover - openpyxl is a declared dependency
        raise AggregationError(
            "MissingDependency",
            "Writing Excel output requires the 'openpyxl' package (see requirements.txt).",
        ) from exc

    keep_macros = ext == ".xlsm"
    if keep_macros and not template_path:
        raise AggregationError(
            "MissingTemplate",
            "Writing a macro-enabled .xlsm requires 'template_path' (an existing .xlsm "
            "whose macros are preserved). openpyxl cannot create VBA from scratch; use a "
            ".xlsx output_path instead if you have no template.",
        )
    if template_path and not os.path.isfile(template_path):
        raise AggregationError("TemplateNotFound", f"Template not found: {template_path}")

    sheet_title = (output_sheet or "").strip() or "Aggregation"
    try:
        if template_path:
            workbook = load_workbook(template_path, keep_vba=keep_macros)
            if sheet_title in workbook.sheetnames:
                worksheet = workbook[sheet_title]
                if worksheet.max_row:
                    worksheet.delete_rows(1, worksheet.max_row)
            else:
                worksheet = workbook.create_sheet(sheet_title)
        else:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = sheet_title

        worksheet.append([str(column) for column in result.columns])
        for row in result.itertuples(index=False, name=None):
            worksheet.append([_excel_safe(value) for value in row])

        directory = os.path.dirname(os.path.abspath(output_path))
        if not os.path.isdir(directory):
            raise AggregationError(
                "WriteError", f"Output directory does not exist: {directory}"
            )
        workbook.save(output_path)
    except AggregationError:
        raise
    except Exception as exc:  # noqa: BLE001 - report write failures as JSON
        raise AggregationError("WriteError", f"Failed to write '{output_path}': {exc}") from exc

    return {
        "path": os.path.abspath(output_path),
        "format": ext.lstrip("."),
        "sheet": sheet_title,
        "row_count": int(len(result)),
        "macros_preserved": keep_macros and bool(template_path),
        "template": os.path.abspath(template_path) if template_path else None,
    }


# --------------------------------------------------------------------------- #
# Public skill
# --------------------------------------------------------------------------- #
def aggregate_data(
    group_by: list,
    metrics: list,
    file_path: str = "",
    data: str = "",
    data_format: str = "auto",
    sheet_name: str = "",
    dataset: str = "",
    sort_by: str = "",
    descending: bool = False,
    limit: int = 0,
    output_path: str = "",
    template_path: str = "",
    output_sheet: str = "",
    include_fingerprint: bool = False,
) -> str:
    """Aggregate rows from an Excel/CSV file or inline data and return grouped results as JSON.

    Use this to compute grouped summaries (totals, averages, counts, distinct counts,
    etc.) over tabular data. Provide the data EITHER via ``file_path`` OR via inline
    ``data``.

    Args:
        group_by: Column names to group by, e.g. ["region", "product"]. Pass an empty
            list [] to aggregate over all rows and return a single grand-total record.
        metrics: Aggregations to compute, each written as "function:column". Examples:
            "sum:amount", "avg:price", "min:qty", "max:qty", "nunique:customer_id".
            Use "count:*" (or "count") for the number of rows in each group. Optionally
            add a custom output name as a third part, e.g. "sum:amount:total_sales".
            Supported functions: sum, mean/avg, min, max, count, nunique/distinct,
            median, std, var, first, last.
        file_path: Path to a .csv, .tsv, .xlsx, .xlsm, .xls, or .json file. Leave empty
            when passing inline ``data``.
        data: Inline data as CSV text or a JSON array of record objects. Leave empty
            when using ``file_path``. If both are given, ``file_path`` is used.
        data_format: One of "auto", "csv", "tsv", "json", "excel". "auto" detects from
            the file extension or inline content. Default "auto".
        sheet_name: For Excel files, the sheet name or zero-based index (as text). Empty
            uses the first sheet.
        dataset: When the input JSON is a multi-table "datasets" envelope (each table
            given as ``headers`` + ``rows``, e.g. another agent's output), selects which
            dataset to aggregate, by name or 0-based index. Empty auto-selects when there
            is only one dataset; with several, the error lists the available names.
            Ignored for CSV/Excel and record-oriented JSON.
        sort_by: Output column to sort by (a group column or a metric output name such
            as "amount_sum"). Empty keeps the natural group order.
        descending: Sort in descending order when True. Default False (ascending).
        limit: Keep only the first N result rows after sorting. 0 (default) keeps all.
        output_path: Optional path ending in .xlsx or .xlsm. When set, the aggregated
            result rows are also written to this Excel file and the JSON response reports
            where it was saved under "output_file". Leave empty to return JSON only.
        template_path: Path to an existing macro-enabled .xlsm workbook. REQUIRED when
            output_path ends in .xlsm: its VBA macros are preserved (openpyxl cannot
            create macros from scratch). Optional for .xlsx output, where a template's
            other sheets are kept. Leave empty to build a fresh .xlsx.
        output_sheet: Name of the worksheet to (re)write with the results. Default
            "Aggregation". Any other sheets in a template are left untouched.
        include_fingerprint: When True, add an "input_fingerprint" object to the response
            for troubleshooting result drift: a row count plus order-insensitive,
            value-canonicalized digests of the columns feeding the aggregation (group-by
            keys and metric source columns). Two runs over the same effective data yield
            identical digests; a differing digest means the input data changed. Default
            False.

    Returns:
        A JSON string. On success:
        {"status": "success", "source": {...}, "rows_read": int, "columns": [...],
         "group_by": [...], "metrics": [...], "record_count": int, "records": [ {...} ]}.
        When output_path is set, an "output_file" object is added:
        {"path": str, "format": "xlsx"|"xlsm", "sheet": str, "row_count": int,
         "macros_preserved": bool, "template": str|null}.
        When include_fingerprint is True, an "input_fingerprint" object is added:
        {"rows": int, "digest": str, "columns": {col: {"digest": str, "non_null": int,
         "null": int}}}.
        On failure:
        {"status": "error", "error_type": str, "message": str}.
    """
    try:
        group_by = _normalize_str_list(group_by)
        metrics = _normalize_str_list(metrics)
        descending = _to_bool(descending)
        limit = _to_int(limit)

        frame, source = _load_dataframe(file_path, data, data_format, sheet_name, dataset)
        if frame.shape[1] == 0:
            raise AggregationError("EmptyData", "The input contains no columns.")

        specs = _parse_metrics(metrics, frame.columns)
        result, warnings, fingerprint = _run_aggregation(
            frame, group_by, specs, sort_by, descending, limit, _to_bool(include_fingerprint)
        )
        output_info = None
        if (output_path or "").strip():
            output_info = _write_results_file(
                result, output_path, template_path, output_sheet
            )
        output = _build_output(
            frame, source, group_by, metrics, result, output_info, warnings, fingerprint
        )
        return json.dumps(output, ensure_ascii=False)
    except AggregationError as err:
        return json.dumps(
            {"status": "error", "error_type": err.error_type, "message": err.message},
            ensure_ascii=False,
        )
    except Exception as exc:  # noqa: BLE001 - never leak a stack trace to the agent
        return json.dumps(
            {"status": "error", "error_type": "UnexpectedError", "message": str(exc)},
            ensure_ascii=False,
        )


def _cli(argv=None) -> int:
    """Command-line entry point. Prints the aggregation result as JSON.

    Exit code 0 on success, 1 when the result has status "error".
    """
    import argparse

    parser = argparse.ArgumentParser(
        prog="aggregate.py",
        description="Aggregate an Excel/CSV file or inline data and print JSON to stdout.",
    )
    parser.add_argument(
        "--group-by",
        default="",
        help='Comma-separated columns to group by, e.g. "region,product". '
        "Omit for a single grand-total record.",
    )
    parser.add_argument(
        "--metrics",
        default="",
        help='Comma-separated "func:column" items, e.g. "sum:amount,count:*,avg:price". '
        'Use "count:*" for the row count.',
    )
    parser.add_argument("--file-path", default="", help="Path to a .csv/.tsv/.xlsx/.xls/.json file.")
    parser.add_argument("--data", default="", help="Inline CSV text or a JSON array of records.")
    parser.add_argument(
        "--data-format",
        default="auto",
        choices=["auto", "csv", "tsv", "json", "excel"],
        help="Input format. Default: auto-detect.",
    )
    parser.add_argument("--sheet-name", default="", help="Excel sheet name or 0-based index.")
    parser.add_argument(
        "--dataset",
        default="",
        help="For a multi-table JSON 'datasets' envelope, the dataset name or 0-based index "
        "to aggregate.",
    )
    parser.add_argument("--sort-by", default="", help="Output column to sort by.")
    parser.add_argument("--descending", action="store_true", help="Sort in descending order.")
    parser.add_argument("--limit", type=int, default=0, help="Keep only the first N result rows.")
    parser.add_argument(
        "--output-path",
        default="",
        help="Optional .xlsx/.xlsm file to also write the results to.",
    )
    parser.add_argument(
        "--template-path",
        default="",
        help="Macro-enabled .xlsm template whose VBA is preserved (required for .xlsm output).",
    )
    parser.add_argument(
        "--output-sheet",
        default="",
        help='Worksheet name to (re)write with the results. Default "Aggregation".',
    )
    parser.add_argument(
        "--include-fingerprint",
        action="store_true",
        help="Add an 'input_fingerprint' block (row count + per-column digests) for "
        "diagnosing result drift between runs.",
    )
    args = parser.parse_args(argv)

    result = aggregate_data(
        group_by=args.group_by,
        metrics=args.metrics,
        file_path=args.file_path,
        data=args.data,
        data_format=args.data_format,
        sheet_name=args.sheet_name,
        dataset=args.dataset,
        sort_by=args.sort_by,
        descending=args.descending,
        limit=args.limit,
        output_path=args.output_path,
        template_path=args.template_path,
        output_sheet=args.output_sheet,
        include_fingerprint=args.include_fingerprint,
    )
    print(result)
    return 0 if json.loads(result).get("status") == "success" else 1


if __name__ == "__main__":
    raise SystemExit(_cli())
